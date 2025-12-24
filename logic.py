from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Tuple, Any, Optional
import os
import re
import openpyxl
import db

# Ratings
RATING_OPTIONS = ["Demonstrated", "Partially Demonstrated", "Not Demonstrated"]

# Evidence Threshold Evaluation Framework thresholds (from your form logic)
THRESHOLD_DEMO = 0.70
THRESHOLD_PARTIAL = 0.40

# ---- weights per TARGET LEVEL (from your provided rule text) ----
# IMPORTANT: This mapping is fixed by your policy, not by Excel.
LEVEL_WEIGHTS: Dict[str, Dict[str, float]] = {
    "Senior Specialist": {
        "Line Manager": 0.10,
        "Second Line Manager": 0.35,
        "OPD": 0.15,
        "HRBP": 0.20,
        "External Stakeholder": 0.20,
    },
    "Lead Expert": {
        "Line Manager": 0.10,
        "Second Line Manager": 0.30,
        "OPD": 0.10,
        "HRBP": 0.20,
        "External Stakeholder 1": 0.15,
        "External Stakeholder 2": 0.15,
    },
    "Advanced Expert": {
        "Line Manager": 0.10,
        "Department Deputy": 0.30,
        "OPD": 0.10,
        "HRBP": 0.20,
        "External Stakeholder 1": 0.15,
        "External Stakeholder 2": 0.15,
    },
    "Principal": {
        "Line Manager": 0.10,
        "Department Deputy": 0.30,
        "HR Deputy": 0.15,
        "HRBP Director": 0.15,
        "External Stakeholder 1": 0.15,
        "External Stakeholder 2": 0.15,
    },
    "Distinguished": {
        "Department Deputy": 0.10,
        "CEO Deputy": 0.30,
        "HR Deputy": 0.15,
        "HRBP Director": 0.15,
        "External Stakeholder 1": 0.15,
        "External Stakeholder 2": 0.15,
    },
}


# ---------- Rules structure loaded from Excel ----------
@dataclass
class Rules:
    dimensions: List[str]                 # A12:A19
    level_paths: List[str]                # B11:F11
    min_demo_by_path: Dict[str, int]      # A25:A29 + B25:B29
    critical_by_path: Dict[str, List[int]]  # 1/0 matrix B12:F19 per path (0-based dim indices)


_ARROW_PAT = re.compile(r"(→|->|➜|⇒)")


def _norm(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _has_arrow(s: str) -> bool:
    return bool(_ARROW_PAT.search(s or ""))


def _persian_to_english_digits(s: str) -> str:
    trans = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
    return (s or "").translate(trans)


def _extract_first_int(s: str) -> Optional[int]:
    if not s:
        return None
    s2 = _persian_to_english_digits(s)
    m = re.search(r"(\d+)", s2)
    return int(m.group(1)) if m else None


def _find_rules_xlsx() -> str:
    """
    No-manual-config: auto-detect an .xlsx in working directory that contains Sheet2.
    Priority:
      1) env RULES_XLSX_PATH (if valid)
      2) ./rules.xlsx (if exists and has Sheet2)
      3) any *.xlsx in cwd that has Sheet2
    """
    env_path = os.environ.get("RULES_XLSX_PATH")
    candidates = []

    if env_path:
        candidates.append(env_path)

    candidates.append("rules.xlsx")

    # add any xlsx in current directory
    for f in os.listdir("."):
        if f.lower().endswith(".xlsx"):
            candidates.append(f)

    for p in candidates:
        if not p or not os.path.exists(p):
            continue
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            if "Sheet2" in wb.sheetnames:
                return p
        except Exception:
            continue

    raise FileNotFoundError(
        "Could not find a rules .xlsx containing 'Sheet2'. "
        "Put your IC Evaluation Excel file in the repo root (same folder as app.py) "
        "or name it rules.xlsx."
    )


def level_path_to_target_level(level_path: str) -> str:
    lp = (level_path or "").replace("->", "→")
    parts = [p.strip() for p in lp.split("→")]
    if len(parts) < 2:
        raise ValueError(f"Invalid level_path: {level_path}")
    return parts[-1]


def load_rules_from_sheet2() -> Rules:
    # Your Sheet2 has known fixed layout:
    # - Critical matrix header row at row 11, col 2..6
    # - Dimensions at col 1 rows 12..19
    # - Critical matrix values at rows 12..19 cols 2..6
    # - Min Demo table at rows 25..29 col 1 is level_path; col 2 contains number
    xlsx_path = _find_rules_xlsx()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Sheet2" not in wb.sheetnames:
        raise ValueError(f"Expected 'Sheet2' in workbook. Found: {wb.sheetnames}")

    ws = wb["Sheet2"]

    # ---- level paths (B11:F11) ----
    level_paths = []
    level_cols = []
    for c in range(2, 7):  # B..F
        v = _norm(ws.cell(11, c).value).replace("->", "→")
        if not v or not _has_arrow(v):
            raise ValueError(f"Sheet2: expected level path with '→' at row 11 col {c}. Found: {v!r}")
        level_paths.append(v)
        level_cols.append(c)

    # ---- dimensions (A12:A19) ----
    dimensions = []
    for r in range(12, 20):  # 12..19
        d = _norm(ws.cell(r, 1).value)
        if not d:
            raise ValueError(f"Sheet2: expected dimension name in A{r} but it was empty.")
        dimensions.append(d)

    if len(dimensions) != 8:
        raise ValueError(f"Sheet2: expected 8 dimensions (A12:A19). Found {len(dimensions)}")

    # ---- critical matrix (B12:F19) ----
    critical_by_path: Dict[str, List[int]] = {p: [] for p in level_paths}
    for i, r in enumerate(range(12, 20)):  # dim index i
        for j, c in enumerate(level_cols):
            raw = ws.cell(r, c).value
            val = None
            if isinstance(raw, (int, float)):
                val = int(raw)
            else:
                t = _norm(raw)
                if t in ("1", "0"):
                    val = int(t)

            if val == 1:
                critical_by_path[level_paths[j]].append(i)

    # ---- min demonstrated table (A25:A29 + B25:B29) ----
    min_demo_by_path: Dict[str, int] = {}
    for r in range(25, 30):
        lp = _norm(ws.cell(r, 1).value).replace("->", "→")
        if not lp:
            continue
        num_cell = _norm(ws.cell(r, 2).value)
        n = _extract_first_int(num_cell)
        if n is None:
            raise ValueError(f"Sheet2: could not extract min demonstrated from B{r}: {num_cell!r}")
        min_demo_by_path[lp] = int(n)

    # ensure all paths exist
    missing = [p for p in level_paths if p not in min_demo_by_path]
    if missing:
        raise ValueError(f"Sheet2: Min Demonstrated missing for paths: {missing}")

    return Rules(
        dimensions=dimensions,
        level_paths=level_paths,
        min_demo_by_path=min_demo_by_path,
        critical_by_path=critical_by_path,
    )


def get_rules() -> Rules:
    if not hasattr(get_rules, "_cache"):
        setattr(get_rules, "_cache", load_rules_from_sheet2())
    return getattr(get_rules, "_cache")


# ---------- Aggregation + decision ----------
def allowed_evaluator_roles_by_level_path(level_path: str) -> List[str]:
    target = level_path_to_target_level(level_path)
    if target not in LEVEL_WEIGHTS:
        raise ValueError(f"Unknown target level derived from level_path: {target}")
    return list(LEVEL_WEIGHTS[target].keys())


def get_weight(level_path: str, evaluator_role: str) -> float:
    target = level_path_to_target_level(level_path)
    return float(LEVEL_WEIGHTS[target].get(evaluator_role, 0.0))


def aggregate_dimension_demo_weight(votes: List[Tuple[str, str]], level_path: str) -> float:
    """
    Evidence Threshold method (as per your Excel):
      - sum ONLY weights where rating == Demonstrated
    """
    w = 0.0
    for role, rating in votes:
        if rating == "Demonstrated":
            w += get_weight(level_path, role)
    return w


def dimension_result_from_demo_weight(demo_weight: float) -> str:
    if demo_weight >= THRESHOLD_DEMO:
        return "Demonstrated"
    if demo_weight >= THRESHOLD_PARTIAL:
        return "Partially Demonstrated"
    return "Not Demonstrated"


def decision_from_dimension_results(results: List[str], level_path: str) -> str:
    rules = get_rules()
    critical = rules.critical_by_path[level_path]
    min_demo = rules.min_demo_by_path[level_path]

    # critical must be Demonstrated
    for idx in critical:
        if results[idx] != "Demonstrated":
            return "Reject"

    demo_count = sum(1 for r in results if r == "Demonstrated")
    return "Confirmed" if demo_count >= int(min_demo) else "Reject"


def committee_aggregate(eval_id: int) -> Dict[str, Any]:
    ev = db.get_evaluation(eval_id)
    if not ev:
        raise ValueError("Evaluation not found")

    level_path = ev["level_path"]
    target_level = ev["target_level"]
    rules = get_rules()

    assignments = db.list_assignments_for_evaluation(eval_id)
    responses = db.list_responses_for_evaluation(eval_id)

    # pending if not all assigned submitted
    all_assigned = {int(a["user_id"]) for a in assignments}
    responded = {int(r["user_id"]) for r in responses}
    if all_assigned and responded != all_assigned:
        return {
            "per_dim": [],
            "committee_decision": "Pending",
            "missing_count": len(all_assigned - responded),
            "responded_count": len(responded),
            "assigned_count": len(all_assigned),
        }

    role_by_user = {int(a["user_id"]): a["evaluator_role"] for a in assignments}

    per_dim_votes: List[List[Tuple[str, str]]] = [[] for _ in range(8)]
    for r in responses:
        uid = int(r["user_id"])
        role = role_by_user.get(uid, "")
        dims = [r["dim1"], r["dim2"], r["dim3"], r["dim4"], r["dim5"], r["dim6"], r["dim7"], r["dim8"]]
        for i in range(8):
            per_dim_votes[i].append((role, dims[i]))

    per_dim_out = []
    dim_results = []
    for i in range(8):
        dw = aggregate_dimension_demo_weight(per_dim_votes[i], level_path)
        res = dimension_result_from_demo_weight(dw)
        per_dim_out.append({"dimension": rules.dimensions[i], "demo_weight": dw, "result": res})
        dim_results.append(res)

    # Committee decision rules:
    # - If target is Principal or Distinguished => RecommendationOnly (final via approver)
    if target_level in ("Principal", "Distinguished"):
        committee_decision = "RecommendationOnly"
    else:
        committee_decision = decision_from_dimension_results(dim_results, level_path)

    return {
        "per_dim": per_dim_out,
        "committee_decision": committee_decision,
        "missing_count": 0,
        "responded_count": len(responses),
        "assigned_count": len(assignments),
    }


def approver_final_decision(eval_id: int) -> Dict[str, Any]:
    ev = db.get_evaluation(eval_id)
    if not ev:
        raise ValueError("Evaluation not found")

    rules = get_rules()
    appr = db.get_approver_response(eval_id)
    if not appr:
        return {"final_decision": "Pending", "per_dim": []}

    dims = [appr["dim1"], appr["dim2"], appr["dim3"], appr["dim4"],
            appr["dim5"], appr["dim6"], appr["dim7"], appr["dim8"]]

    final = decision_from_dimension_results(dims, ev["level_path"])
    per_dim = [{"dimension": rules.dimensions[i], "result": dims[i]} for i in range(8)]
    return {"final_decision": final, "per_dim": per_dim}
